#!/usr/bin/env python3
"""
Validate Excel inventory and produce a structured validation report.

This script reuses reference JSONs in `DE_PYTHON_NUEVO` (resistors, transistors, integrated_circuits,
capacitors_ceramic) to detect known parts and to flag unknowns. It also detects column types,
multi-value cells, duplicated rows and formula cells. Outputs JSON and CSV reports in `reports/`.

Usage:
  python scripts/ingest/validate_excel.py ./Inventario_Cirujanosintetizadores.xlsx --output reports/validation_report.json

"""
from __future__ import annotations
import argparse
import json
import os
import hashlib
from collections import defaultdict
from datetime import datetime
import csv

import pandas as pd
from dateutil import parser as dateparser
from openpyxl import load_workbook

# Paths to reference JSONs
REF_DIR = os.path.join(os.path.dirname(__file__), '..', '..', 'DE_PYTHON_NUEVO')
REF_DIR = os.path.normpath(REF_DIR)
REF_RESISTORS = os.path.join(REF_DIR, 'resistors.json')
REF_CAPS = os.path.join(REF_DIR, 'capacitors_ceramic.json')
REF_ICS = os.path.join(REF_DIR, 'integrated_circuits.json')
REF_TRANS = os.path.join(REF_DIR, 'transistors.json')


def md5_of_file(path, blocksize=65536):
    h = hashlib.md5()
    with open(path, 'rb') as f:
        for block in iter(lambda: f.read(blocksize), b''):
            h.update(block)
    return h.hexdigest()


def load_references():
    refs = {}
    for k, p in [('resistors', REF_RESISTORS), ('capacitors', REF_CAPS), ('ics', REF_ICS), ('transistors', REF_TRANS)]:
        try:
            with open(p, 'r', encoding='utf-8') as fh:
                refs[k] = json.load(fh)
        except Exception:
            refs[k] = []
    # build quick lookups
    lookups = {
        'transistors': {r.get('part_number', '').upper(): r for r in refs.get('transistors', []) if r.get('part_number')},
        'ics': {r.get('part_number', '').upper(): r for r in refs.get('ics', []) if r.get('part_number')},
    }
    return refs, lookups


def detect_value_types(values):
    types = set()
    examples = defaultdict(list)
    for v in values:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        s = str(v).strip()
        if s == '':
            continue
        ls = s.lower()
        if ls in ('true', 'false', 'si', 'no'):
            types.add('bool')
            examples['bool'].append(s)
            continue
        try:
            if '.' not in s and s.replace('-', '').isdigit():
                int(s)
                types.add('integer')
                examples['integer'].append(s)
                continue
        except Exception:
            pass
        try:
            float(s.replace(',', '.'))
            types.add('float')
            examples['float'].append(s)
            continue
        except Exception:
            pass
        try:
            _ = dateparser.parse(s, dayfirst=True)
            types.add('date')
            examples['date'].append(s)
            continue
        except Exception:
            pass
        types.add('text')
        examples['text'].append(s)
    return list(types), {k: examples[k][:5] for k in examples}


def detect_multivalue(values):
    delimiters = [',', ';', '/', '|']
    count = 0
    examples = []
    for v in values:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        s = str(v)
        for d in delimiters:
            if d in s and len(s.split(d)) > 1:
                count += 1
                if len(examples) < 5:
                    examples.append(s)
                break
    return {'multi_value_flag': count > 0, 'samples': examples, 'count': count}


def detect_formulas(path, sheet_name):
    try:
        wb = load_workbook(filename=path, read_only=True, data_only=False)
    except Exception:
        return {}
    if isinstance(sheet_name, int):
        sheet_name = wb.sheetnames[sheet_name]
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    col_formulas = defaultdict(list)
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            try:
                if cell.data_type == 'f' or (isinstance(cell.value, str) and str(cell.value).startswith('=')):
                    col_formulas[cell.column_letter].append(cell.coordinate)
            except Exception:
                pass
    return col_formulas


def idx_to_letter(i):
    i += 1
    string = ''
    while i > 0:
        i, remainder = divmod(i-1, 26)
        string = chr(65 + remainder) + string
    return string


def build_report(df, excel_path, lookups, sheet_name=0):
    report = {
        'file': os.path.abspath(excel_path),
        'file_md5': md5_of_file(excel_path),
        'scanned_at': datetime.utcnow().isoformat() + 'Z',
        'row_count': int(df.shape[0]),
        'column_count': int(df.shape[1]),
        'columns': [],
        'duplicates_summary': {},
        'issues': [],
        'reference_matches': {}
    }

    try:
        col_formulas_map = detect_formulas(excel_path, sheet_name)
    except Exception:
        col_formulas_map = {}

    for col in df.columns:
        series = df[col]
        non_null = series.dropna()
        non_null_count = int(non_null.shape[0])
        null_count = int(df.shape[0] - non_null_count)
        unique_vals = non_null.drop_duplicates().head(10).tolist()

        detected_types, examples = detect_value_types(non_null.tolist())
        multiv = detect_multivalue(non_null.tolist())

        col_idx = df.columns.get_loc(col)
        col_letter = idx_to_letter(col_idx)
        formulas = col_formulas_map.get(col_letter, []) if col_formulas_map else []

        # reference matching for part-like columns
        matches = {'ics': 0, 'transistors': 0, 'resistors': 0}
        sample_unmatched = []
        for v in non_null.head(200):
            s = str(v).strip()
            if not s:
                continue
            up = s.upper()
            if up in lookups.get('ics', {}):
                matches['ics'] += 1
                continue
            if up in lookups.get('transistors', {}):
                matches['transistors'] += 1
                continue
            try:
                fv = float(s)
                matches['resistors'] += 1
                continue
            except Exception:
                sample_unmatched.append(s)
                if len(sample_unmatched) > 10:
                    break

        col_report = {
            'name': str(col),
            'index': int(col_idx),
            'non_null_count': non_null_count,
            'null_count': null_count,
            'unique_values_sample': unique_vals,
            'detected_types': detected_types,
            'detected_type_examples': examples,
            'multi_value': multiv['multi_value_flag'],
            'multi_value_examples': multiv['samples'],
            'formula_cells': formulas,
            'reference_matches': matches,
            'unmatched_sample': sample_unmatched[:5]
        }
        report['columns'].append(col_report)

    # duplicates
    try:
        dup_mask = df.duplicated(keep=False)
        dup_count = int(dup_mask.sum())
        dup_examples = df[dup_mask].head(10).to_dict(orient='records')
    except Exception:
        dup_count = 0
        dup_examples = []
    report['duplicates_summary']['exact_duplicates_count'] = dup_count
    report['duplicates_summary']['exact_duplicates_examples'] = dup_examples

    # candidate id columns
    candidate_id_names = ['sku', 'codigo', 'code', 'part_number', 'part no', 'partnumber', 'id', 'item_id']
    found_candidates = [c for c in df.columns if str(c).strip().lower() in candidate_id_names]
    candidate_dup_info = {}
    if found_candidates:
        for cand in found_candidates:
            try:
                cdup_mask = df.duplicated(subset=[cand], keep=False)
                cand_dup_count = int(cdup_mask.sum())
                candidate_dup_info[str(cand)] = {
                    'duplicates_count': cand_dup_count,
                    'examples': df[cdup_mask].head(5).to_dict(orient='records')
                }
            except Exception:
                candidate_dup_info[str(cand)] = {'duplicates_count': 0, 'examples': []}
    report['duplicates_summary']['candidate_key_duplicates'] = candidate_dup_info

    # issues
    for c in report['columns']:
        if len(c['detected_types']) > 1:
            report['issues'].append({
                'type': 'mixed_types',
                'column': c['name'],
                'detected_types': c['detected_types'],
                'examples': c.get('detected_type_examples', {})
            })
        if c['null_count'] / max(1, df.shape[0]) > 0.5:
            report['issues'].append({
                'type': 'high_null_ratio',
                'column': c['name'],
                'null_count': c['null_count'],
                'row_count': df.shape[0]
            })
        if c['multi_value']:
            report['issues'].append({
                'type': 'multi_value_cells',
                'column': c['name'],
                'examples': c['multi_value_examples']
            })
        if c['formula_cells']:
            report['issues'].append({
                'type': 'formula_cells',
                'column': c['name'],
                'examples': c['formula_cells'][:5]
            })

    return report


def save_reports(report_json_path, report):
    os.makedirs(os.path.dirname(report_json_path) or '.', exist_ok=True)
    with open(report_json_path, 'w', encoding='utf-8') as fh:
        json.dump(report, fh, ensure_ascii=False, indent=2)
    # CSV summary
    out_csv = os.path.splitext(report_json_path)[0] + '.csv'
    with open(out_csv, 'w', newline='', encoding='utf-8') as csvf:
        writer = csv.writer(csvf)
        writer.writerow(['column','index','non_null_count','null_count','detected_types','multi_value','formula_cells','matched_ics','matched_transistors'])
        for c in report['columns']:
            matched_ics = c.get('reference_matches', {}).get('ics', 0)
            matched_trans = c.get('reference_matches', {}).get('transistors', 0)
            writer.writerow([c['name'], c['index'], c['non_null_count'], c['null_count'], '|'.join(c['detected_types']), c['multi_value'], len(c['formula_cells']) if c.get('formula_cells') else 0, matched_ics, matched_trans])
    print(f'Reports written: {report_json_path} and {out_csv}')


def main():
    p = argparse.ArgumentParser(description='Validate Excel inventory and produce reports')
    p.add_argument('excel', nargs='?', default='./Inventario_Cirujanosintetizadores.xlsx', help='Excel file path')
    p.add_argument('--sheet', default=0, help='Sheet index (int) or name', type=str)
    p.add_argument('--output', default='reports/validation_report.json', help='Output JSON path')
    args = p.parse_args()

    excel_path = args.excel
    if not os.path.exists(excel_path):
        raise SystemExit(f'Excel file not found: {excel_path}')

    refs, lookups = load_references()

    try:
        sheet = int(args.sheet)
    except Exception:
        sheet = args.sheet

    df = pd.read_excel(excel_path, sheet_name=sheet, dtype=object, engine='openpyxl')

    report = build_report(df, excel_path, lookups, sheet_name=sheet)
    save_reports(args.output, report)


if __name__ == '__main__':
    main()
